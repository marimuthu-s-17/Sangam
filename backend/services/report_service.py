from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from models.expense import Expense
from models.loan import Loan, LoanPayment
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
import csv


class ReportService:
    """Service layer for fetching report data and generating CSV, Excel, and PDF exports."""

    @staticmethod
    async def get_monthly_expense_data(db: AsyncSession, year: int, month: int) -> list[dict]:
        """Fetch all expenses within a given year and month."""
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        stmt = (
            select(Expense)
            .where(and_(Expense.expense_date >= start_date, Expense.expense_date < end_date, Expense.is_deleted == False))
            .order_by(Expense.expense_date.asc())
        )
        res = await db.execute(stmt)
        expenses = res.scalars().all()
        return [
            {
                "id": e.id,
                "date": str(e.expense_date),
                "category": e.category,
                "description": e.description,
                "amount": float(e.amount),
                "payment_method": e.payment_method,
                "remarks": e.remarks or e.notes or "",
            }
            for e in expenses
        ]

    @staticmethod
    async def get_category_expense_data(db: AsyncSession, category: str) -> list[dict]:
        """Fetch all expenses under a specific category."""
        stmt = (
            select(Expense)
            .where(Expense.category == category, Expense.is_deleted == False)
            .order_by(Expense.expense_date.desc())
        )
        res = await db.execute(stmt)
        expenses = res.scalars().all()
        return [
            {
                "id": e.id,
                "date": str(e.expense_date),
                "category": e.category,
                "description": e.description,
                "amount": float(e.amount),
                "payment_method": e.payment_method,
                "remarks": e.remarks or e.notes or "",
            }
            for e in expenses
        ]

    @staticmethod
    async def get_loan_data(db: AsyncSession) -> list[dict]:
        """Fetch all loans with dynamic interest and outstanding recalculations."""
        stmt = select(Loan).options(selectinload(Loan.payments)).order_by(Loan.loan_date.desc())
        res = await db.execute(stmt)
        loans = res.scalars().all()

        from services.loan_service import LoanService

        data = []
        for loan in loans:
            LoanService.recalculate_loan(loan)
            data.append({
                "id": loan.id,
                "borrower_name": loan.borrower_name,
                "phone_number": loan.phone_number,
                "loan_amount": float(loan.loan_amount),
                "interest_rate": float(loan.interest_rate),
                "loan_date": str(loan.loan_date),
                "due_date": str(loan.due_date),
                "outstanding_amount": float(loan.outstanding_amount),
                "status": loan.status,
                "remarks": loan.remarks or "",
            })
        await db.flush()
        return data

    @staticmethod
    async def get_interest_collection_data(db: AsyncSession) -> list[dict]:
        """Fetch all recorded interest payments."""
        stmt = (
            select(LoanPayment, Loan)
            .join(Loan, Loan.id == LoanPayment.loan_id)
            .where(LoanPayment.interest_payment > 0)
            .order_by(LoanPayment.payment_date.desc())
        )
        res = await db.execute(stmt)
        rows = res.all()
        return [
            {
                "id": payment.id,
                "payment_id": payment.id,
                "loan_id": loan.id,
                "borrower_name": loan.borrower_name,
                "payment_date": str(payment.payment_date),
                "interest_payment": float(payment.interest_payment),
                "payment_method": payment.payment_method,
                "notes": payment.notes or "",
            }
            for payment, loan in rows
        ]

    @staticmethod
    async def get_member_report_data(db: AsyncSession) -> list[dict]:
        """Fetch member registration details and summaries."""
        from models.member import Member
        from models.monthly_auction import MonthlyContribution
        
        stmt = select(Member).order_by(Member.name)
        res = await db.execute(stmt)
        members = res.scalars().all()
        
        data = []
        for member in members:
            # Contributions
            contributions_stmt = select(
                func.coalesce(func.sum(MonthlyContribution.paid_amount), 0),
                func.coalesce(
                    func.sum(
                        func.greatest(
                            MonthlyContribution.minimum_amount
                            - MonthlyContribution.paid_amount,
                            0,
                        )
                    ),
                    0,
                ),
            ).where(MonthlyContribution.member_id == member.id)
            contrib_res = await db.execute(contributions_stmt)
            total_contrib, outstanding = contrib_res.first()
            
            data.append({
                "id": member.id,
                "name": member.name,
                "phone": member.phone,
                "joined_date": str(member.created_at.date()) if member.created_at else "",
                "status": member.status,
                "total_contributions": float(total_contrib),
                "outstanding_balance": float(outstanding)
            })
        await db.flush()
        return data

    @staticmethod
    async def get_auction_report_data(db: AsyncSession) -> list[dict]:
        """Fetch all auctions with their status, current round, and prize information."""
        from models.auction import Auction
        
        stmt = select(Auction).order_by(Auction.id)
        res = await db.execute(stmt)
        auctions = res.scalars().all()
        
        return [
            {
                "id": a.id,
                "name": a.name,
                "start_month": str(a.start_month) if a.start_month else "",
                "duration": a.total_months,
                "prize_amount": float(a.prize_amount),
                "status": a.status,
                "current_month": a.current_month
            }
            for a in auctions
        ]

    @staticmethod
    async def get_profit_loss_report_data(db: AsyncSession) -> list[dict]:
        """Generate Profit & Loss summary metrics."""
        from models.monthly_auction import MonthlyContribution
        from models.finance import Finance, TransactionType
        from models.expense import Expense
        from models.loan import Loan, LoanPayment
        
        total_contrib = (await db.execute(select(func.coalesce(func.sum(MonthlyContribution.paid_amount), 0)))).scalar() or Decimal("0.00")
        total_receipts = (await db.execute(select(func.coalesce(func.sum(Finance.amount), 0)).where(Finance.transaction_type == TransactionType.RECEIPT))).scalar() or Decimal("0.00")
        total_payments = (await db.execute(select(func.coalesce(func.sum(Finance.amount), 0)).where(Finance.transaction_type == TransactionType.PAYMENT))).scalar() or Decimal("0.00")
        
        total_expenses = (await db.execute(select(func.coalesce(func.sum(Expense.amount), 0)).where(Expense.is_deleted == False))).scalar() or Decimal("0.00")
        total_loans_amount = (await db.execute(select(func.coalesce(func.sum(Loan.loan_amount), 0)))).scalar() or Decimal("0.00")
        outstanding_loans = (await db.execute(select(func.coalesce(func.sum(Loan.outstanding_amount), 0)).where(Loan.status != "closed"))).scalar() or Decimal("0.00")
        interest_earned = (await db.execute(select(func.coalesce(func.sum(LoanPayment.interest_payment), 0)))).scalar() or Decimal("0.00")
        loan_principal_repayments = (await db.execute(select(func.coalesce(func.sum(LoanPayment.principal_payment), 0)))).scalar() or Decimal("0.00")
        
        cash_inflows = total_contrib + total_receipts + interest_earned + loan_principal_repayments
        cash_outflows = total_expenses + total_loans_amount + total_payments
        available_balance = cash_inflows - cash_outflows
        
        overall_profit_loss = interest_earned + total_receipts - total_expenses - total_payments
        
        return [
            {"id": 1, "metric": "Total Collections (Contributions)", "amount": float(total_contrib)},
            {"id": 2, "metric": "General Receipts", "amount": float(total_receipts)},
            {"id": 3, "metric": "Total Expenses", "amount": float(total_expenses)},
            {"id": 4, "metric": "General Payments", "amount": float(total_payments)},
            {"id": 5, "metric": "Interest Earned from Loans", "amount": float(interest_earned)},
            {"id": 6, "metric": "Outstanding Loan Principal", "amount": float(outstanding_loans)},
            {"id": 7, "metric": "Cash Available (Sangam Balance)", "amount": float(available_balance)},
            {"id": 8, "metric": "Overall Net Profit / Loss", "amount": float(overall_profit_loss)}
        ]

    @staticmethod
    async def get_outstanding_loan_data(db: AsyncSession) -> list[dict]:
        """Fetch all loans with a remaining outstanding balance > 0."""
        stmt = select(Loan).options(selectinload(Loan.payments)).order_by(Loan.loan_date.desc())
        res = await db.execute(stmt)
        loans = res.scalars().all()

        from services.loan_service import LoanService

        data = []
        for loan in loans:
            LoanService.recalculate_loan(loan)
            if loan.outstanding_amount > 0:
                data.append({
                    "id": loan.id,
                    "borrower_name": loan.borrower_name,
                    "phone_number": loan.phone_number,
                    "loan_amount": float(loan.loan_amount),
                    "interest_rate": float(loan.interest_rate),
                    "loan_date": str(loan.loan_date),
                    "due_date": str(loan.due_date),
                    "outstanding_amount": float(loan.outstanding_amount),
                    "status": loan.status,
                    "remarks": loan.remarks or "",
                })
        await db.flush()
        return data

    @staticmethod
    def generate_csv(headers: list[str], rows: list[list]) -> bytes:
        """Helper to generate a raw CSV binary file."""
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")

    @staticmethod
    def generate_excel(title: str, headers: list[str], rows: list[list]) -> bytes:
        """Helper to generate a beautifully styled Excel binary file using openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border_side = Side(border_style="thin", color="D9D9D9")
        data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Title
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Meta-info
        meta_str = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Report Period: Cumulative"
        ws.append([meta_str])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1).font = meta_font
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        ws.append([])  # Space row

        # Headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row in rows:
            ws.append(row)

        # Style data cells
        total_row_idx = len(rows) + 4
        for r_idx in range(5, total_row_idx + 1):
            is_total_row = (r_idx == total_row_idx and ws.cell(row=r_idx, column=1).value and "Total" in str(ws.cell(row=r_idx, column=1).value))
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = data_border
                
                if is_total_row:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        file_stream = BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def generate_pdf(title: str, headers: list[str], rows: list[list]) -> bytes:
        """Helper to generate a professional PDF document using reportlab."""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from datetime import datetime

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1A237E"),
            alignment=1,  # Centered
            spaceAfter=5,
        )
        story.append(Paragraph(title, title_style))

        # Meta-info
        meta_style = ParagraphStyle(
            name="ReportMeta",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=10,
            textColor=colors.HexColor("#555555"),
            alignment=1,  # Centered
            spaceAfter=15,
        )
        meta_str = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Report Period: Cumulative"
        story.append(Paragraph(meta_str, meta_style))
        story.append(Spacer(1, 5))

        page_width = 612 - 72  # 540 pt
        col_width = page_width / len(headers)
        col_widths = [col_width] * len(headers)

        cell_style = ParagraphStyle(
            name="TableCell", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10
        )
        header_style = ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
        )

        formatted_table_data = []
        # Header row
        formatted_table_data.append([Paragraph(str(h), header_style) for h in headers])
        
        # Data rows
        for idx, r in enumerate(rows):
            is_last = (idx == len(rows) - 1)
            is_total = is_last and isinstance(r[0], str) and "Total" in r[0]
            
            row_cells = []
            for cell in r:
                text = str(cell)
                if is_total:
                    c_style = ParagraphStyle(
                        name=f"TableCellTotal_{idx}", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10
                    )
                else:
                    c_style = cell_style
                row_cells.append(Paragraph(text, c_style))
            formatted_table_data.append(row_cells)

        t = Table(formatted_table_data, colWidths=col_widths)
        
        t_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A237E")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F9F9F9")],
            ),
        ]

        if len(rows) > 0 and isinstance(rows[-1][0], str) and "Total" in rows[-1][0]:
            t_style.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EAEAEA")))
            t_style.append(("LINEABOVE", (0, -1), (-1, -1), 1.5, colors.HexColor("#1A237E")))

        t.setStyle(TableStyle(t_style))
        story.append(t)
        doc.build(story)

        buffer.seek(0)
        return buffer.getvalue()
